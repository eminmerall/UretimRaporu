from  openpyxl import Workbook
#Excel'de kaydetme için kütüphane çağırma.

liste=[]
urunTip=[]
toplamUrun=0
enMak=0
enMakEx=0
urunTipEx=0
tipUrunEx=0
makSay=1
makSayEx=0
kitap = Workbook()                  # Kitap ismine Workbook() değerini atama.
kitap.save("dosya.xlsx")            # Excel dosyasını kaydetme.
kitap.create_sheet("Sayfa1")        # Excel'de Sayfa1 adında sayfa oluşturma.
Sayfa1 = kitap.active               # Aktif sayfaya Sayfa1 adını verme.



while True:             # 2 değer için de sadece sayı girilene kadar döngü
    try:                                            # Sayı dışındaki karakterlerin kontrolü
        makine=int(input("Makine sayısını giriniz: "))
        if(makine<=0):
            print("Makine sayısı en az 1 olmalıdır!")
            continue
        makine=abs(makine)
    except ValueError as hata:                      # Sayı dışında karakter girilirse yapılacaklar.
        print("Lütfen makine adedi için sadece doğal sayı girin!")
        continue
    
    try:                                             # Sayı dışındaki karakterlerin kontrolü
        urun=int(input("Ürün sayısını giriniz: "))
        if(urun<=0):
            print("Ürün sayısı en az 1 olmalıdır!")
            continue
    except ValueError as hata:                       # Sayı dışında karakter girilirse yapılacaklar.
        print("Lütfen ürün adedi için sadece doğal sayı girin!")
        continue


    # Listeyi(matrisi) oluşturma.
    for i in range(urun):
        urunTip.append(0)
    for i in range(makine):
        liste += [[i] *urun]

                    #Makine sayısı kadar döngü   
    for i in range(makine):
        print("%a. makinenin ürettiği"%(i+1))
                    #Ürün Sayısı kadar döngü
        for j in range(urun):
            while True:                             # Sayı dışındaki karakterlerin kontrolü
                try:
                    sayi=int(input("          %a. ürünün adedi: "%(j+1)))
                    if(sayi<0):
                        print("Ürün adedi pozitif olmalıdır!")
                    else:
                        urunTip[j] +=sayi
                        liste[i][j]=sayi
                        toplamUrun +=sayi
                        break
                except ValueError as hata:          # Sayı dışında karakter girilirse yapılacaklar.
                    print("lütfen sadece doğal sayı girin!")
                    continue



        
    print("----------------------------------------------")

                
    print("A: İşletmede üretilen toplam ürün miktarı: ", toplamUrun)        #İşletmede üretilen toplam ürün miktarı.
                #--------- Excel'e A: yi kaydetme
    Sayfa1["C2"]="A:"
    Sayfa1["D2"]=toplamUrun

    print("-----")
    #her ürünün toplam üretim miktarı
    if(urun==0):    #Ürün 0 girilirse ürün üretilmedi yazdırır 
        print("B: Ürün girilmedi!")
    else:
       print("B:")
       for i in range(urun):
        print("İşletmede üretilen",i+1,". ürün miktarı : ",urunTip[i])
        tipUrunEx=urunTip[i]
                #Excel'e B: yi kaydetme
        Sayfa1["C3"]="B:"
        Sayfa1["D3"]=tipUrunEx

    #x makinesindeki toplam ürün miktarı
    print("-----")
    if(makine==0):
        print("C: Makine girilmedi!")
    else:
        print("C:")
        for a in liste:
           print(makSay, ".Makinedeki toplam üretim sayısı:", sum(a))
           makSay += 1
           makSayEx=sum(a)
                        # Excel'e C: yi kaydetme
        Sayfa1["C4"]="C:"
        Sayfa1["D4"]=makSayEx


    print("-----")

    #En fazla üretim yapan makinenin toplam üretim sayısı
    if (makine== 0 or urun == 0):    ##Ürün veya makine 0 girilirse ürün üretilmedi yazdırır 
        print("D: En fazla üretim yapan makinenin ürettiği ürün sayısı: Makineler üretim yapmadı!")
        print("-----")
    else:
        enMak=max(liste)    
        print("D: en fazla üretim yapan makinenin ürettiği ürün sayısı: ", sum(enMak))
        enMakEx=sum(enMak)
        print("-----")
                    # Excel'e D: yi kaydetme
        Sayfa1["C5"]="D:"
        Sayfa1["D5"]=enMakEx

        
        #En fazla üretilen ürün miktarı
    if (makine==0 or urun == 0):    #Ürün veya makine 0 girilirse ürün üretilmedi yazdırır 
        print("en fazla üretilen ürünün miktarı: Ürün üretilmedi!")
        print("")
    else: 
        print("E: en fazla üretilen ürünün miktarı: ",max(urunTip))
        urunTipEx=max(urunTip)
        print("")
                    # Excel'e E: yi kaydetme
        Sayfa1["C6"]="E:"
        Sayfa1["D6"]=urunTipEx
    break
kitap.save("dosya.xlsx")
kitap.close()
